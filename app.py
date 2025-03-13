import streamlit as st
from dotenv import load_dotenv
import os
import tempfile
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from io import StringIO
import docx2txt
import openpyxl
import csv
import ebooklib
from ebooklib import epub
from bs4 import BeautifulSoup
import subprocess
import re
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
# Update the import to use langchain_community instead of langchain
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain.vectorstores import FAISS
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
# Update these imports as well
from langchain_community.llms import HuggingFaceHub
from langchain_community.llms import Ollama
from langchain.chains.summarize import load_summarize_chain
from langchain.docstore.document import Document
from htmlview import css, bot_template, user_template  # Updated import statement

# Add these imports at the top of the file
import json
import requests
import subprocess
from pathlib import Path
import re
# Add this import at the top
from langchain_openai import ChatOpenAI

# Function to extract text from PDF
def get_pdf_text(pdf_file):
    text = ""
    pdf_reader = PdfReader(pdf_file)
    for page in pdf_reader.pages:
        text += page.extract_text()
    return text

# Function to extract text from DOCX
def get_docx_text(docx_file):
    text = docx2txt.process(docx_file)
    return text

# Function to extract text from CSV
def get_csv_text(csv_file):
    text = ""
    csv_content = csv_file.read().decode('utf-8')
    reader = csv.reader(StringIO(csv_content))
    for row in reader:
        text += " | ".join(row) + "\n"
    return text, csv_content

# Function to extract text from Excel
def get_excel_text(excel_file):
    text = ""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(excel_file.getvalue())
        tmp_path = tmp.name
    
    workbook = openpyxl.load_workbook(tmp_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        text += f"Sheet: {sheet}\n"
        for row in worksheet.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
            text += row_text + "\n"
    
    os.unlink(tmp_path)  # Delete the temp file
    return text

# Function to extract text from EPUB
def get_epub_text(epub_file):
    text = ""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.epub') as tmp:
        tmp.write(epub_file.getvalue())
        tmp_path = tmp.name
    
    try:
        book = epub.read_epub(tmp_path)
        
        # Extract metadata
        if book.get_metadata('DC', 'title'):
            text += f"Title: {book.get_metadata('DC', 'title')[0][0]}\n"
        if book.get_metadata('DC', 'creator'):
            text += f"Author: {book.get_metadata('DC', 'creator')[0][0]}\n"
        
        # Extract content
        for item in book.get_items():
            if item.get_type() == ebooklib.ITEM_DOCUMENT:
                content = item.get_content().decode('utf-8')
                soup = BeautifulSoup(content, 'html.parser')
                chapter_text = soup.get_text()
                text += chapter_text + "\n\n"
    except Exception as e:
        st.error(f"Error processing EPUB file: {e}")
    finally:
        os.unlink(tmp_path)  # Delete the temp file
    
    return text

# Function to extract text from DJVU
def get_djvu_text(djvu_file):
    text = ""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.djvu') as tmp:
        tmp.write(djvu_file.getvalue())
        tmp_path = tmp.name
    
    try:
        # Check if djvutxt is installed
        try:
            # Use djvutxt command-line tool to extract text
            result = subprocess.run(['djvutxt', tmp_path, '-'], 
                                   stdout=subprocess.PIPE, 
                                   stderr=subprocess.PIPE,
                                   text=True,
                                   check=True)
            text = result.stdout
        except FileNotFoundError:
            st.error("DjVuLibre tools not found. Please install DjVuLibre to process DJVU files.")
            st.info("On Ubuntu/Debian: sudo apt-get install djvulibre-bin")
            st.info("On macOS: brew install djvulibre")
            st.info("On Windows: Download from http://djvu.sourceforge.net/")
        except subprocess.CalledProcessError as e:
            st.error(f"Error processing DJVU file: {e}")
    finally:
        os.unlink(tmp_path)  # Delete the temp file
    
    return text

# Function to process all document types
def get_document_text(docs):
    text = ""
    dataframes = {}
    
    for doc in docs:
        try:
            if doc.name.endswith('.pdf'):
                st.info(f"Processing PDF: {doc.name}")
                text += get_pdf_text(doc)
            elif doc.name.endswith('.docx'):
                st.info(f"Processing DOCX: {doc.name}")
                text += get_docx_text(doc)
            elif doc.name.endswith('.csv'):
                st.info(f"Processing CSV: {doc.name}")
                csv_text, csv_content = get_csv_text(doc)
                text += csv_text
                # Store dataframe for visualization
                dataframes[doc.name] = pd.read_csv(StringIO(csv_content))
            elif doc.name.endswith(('.xlsx', '.xls')):
                st.info(f"Processing Excel: {doc.name}")
                text += get_excel_text(doc)
                # Store dataframe for visualization
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(doc.getvalue())
                    tmp_path = tmp.name
                dataframes[doc.name] = pd.read_excel(tmp_path)
                os.unlink(tmp_path)
            elif doc.name.endswith('.epub'):
                st.info(f"Processing EPUB: {doc.name}")
                text += get_epub_text(doc)
            elif doc.name.endswith('.djvu'):
                st.info(f"Processing DJVU: {doc.name}")
                text += get_djvu_text(doc)
            else:
                st.warning(f"Unsupported file type: {doc.name}")
        except Exception as e:
            st.error(f"Error processing {doc.name}: {str(e)}")
    
    return text, dataframes

def get_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size=1000,
        chunk_overlap=200,
        length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks

def get_vectorstore(text_chunks):
    st.info("Loading embedding model for vector search (separate from your selected LLM)...")
    
    try:
        # Try using a smaller, more efficient model first
        from langchain_community.embeddings import HuggingFaceEmbeddings
        
        st.info("Loading all-MiniLM-L6-v2 embedding model (for vector search only)...")
        embeddings = HuggingFaceEmbeddings(
            model_name="all-MiniLM-L6-v2",
            model_kwargs={'device': 'cpu'}
        )
        st.success("Embeddings model loaded successfully!")
        
    except Exception as e:
        st.error(f"Error loading primary embedding model: {str(e)}")
        st.info("Trying fallback to basic HuggingFace embeddings...")
        
        try:
            # Even more basic fallback
            from langchain_community.embeddings import HuggingFaceEmbeddings
            embeddings = HuggingFaceEmbeddings()
            st.success("Fallback embeddings model loaded successfully!")
        except Exception as e2:
            st.error(f"Error loading fallback embedding model: {str(e2)}")
            st.error("Could not load any embedding models. Please check your internet connection and try again.")
            st.stop()
    
    # Create and return the vector store
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore

def get_conversation_chain(vectorstore, use_ollama=False, use_openai=False, model_name=None):
    if use_ollama:
        try:
            # Use the selected Ollama model or default to llama2
            ollama_model = model_name if model_name else "llama2"
            llm = Ollama(model=ollama_model)
        except Exception as e:
            st.error(f"Error connecting to Ollama: {e}")
            st.info("Falling back to Hugging Face model")
            llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})
    elif use_openai:
        try:
            # Check if API key is available
            if not os.getenv("OPENAI_API_KEY"):
                st.error("OpenAI API key not found. Please enter your API key in the sidebar.")
                st.stop()
                
            # Use the selected OpenAI model or default to gpt-3.5-turbo
            openai_model = model_name if model_name else "gpt-3.5-turbo"
            llm = ChatOpenAI(model_name=openai_model, temperature=0.5)
        except Exception as e:
            st.error(f"Error connecting to OpenAI: {e}")
            st.info("Falling back to Hugging Face model")
            llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})
    else:
        # Use the selected Hugging Face model or default to flan-t5-xxl
        hf_model = model_name if model_name else "google/flan-t5-xxl"
        llm = HuggingFaceHub(repo_id=hf_model, model_kwargs={"temperature":0.5, "max_length":512})
    
    memory = ConversationBufferMemory(memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain

def generate_summary(text_chunks, use_ollama=False, use_openai=False, model_name=None):
    if use_ollama:
        try:
            # Use the selected Ollama model or default to llama2
            ollama_model = model_name if model_name else "llama2"
            llm = Ollama(model=ollama_model)
        except Exception as e:
            st.error(f"Error connecting to Ollama: {e}")
            st.info("Falling back to Hugging Face model")
            llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})
    elif use_openai:
        try:
            # Check if API key is available
            if not os.getenv("OPENAI_API_KEY"):
                st.error("OpenAI API key not found. Please enter your API key in the sidebar.")
                st.stop()
                
            # Use the selected OpenAI model or default to gpt-3.5-turbo
            openai_model = model_name if model_name else "gpt-3.5-turbo"
            llm = ChatOpenAI(model_name=openai_model, temperature=0.5)
        except Exception as e:
            st.error(f"Error connecting to OpenAI: {e}")
            st.info("Falling back to Hugging Face model")
            llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})
    else:
        # Use the selected Hugging Face model or default to flan-t5-xxl
        hf_model = model_name if model_name else "google/flan-t5-xxl"
        llm = HuggingFaceHub(repo_id=hf_model, model_kwargs={"temperature":0.5, "max_length":512})
    
    docs = [Document(page_content=chunk) for chunk in text_chunks]
    chain = load_summarize_chain(llm, chain_type="map_reduce")
    summary = chain.run(docs)
    return summary

def visualize_data(dataframes):
    if not dataframes:
        st.warning("No data files (CSV/Excel) were uploaded for visualization.")
        return
    
    st.subheader("Data Visualization")
    
    # Select file to visualize
    file_to_viz = st.selectbox("Select a file to visualize:", list(dataframes.keys()))
    df = dataframes[file_to_viz]
    
    # Display the dataframe
    st.write("Data Preview:")
    st.dataframe(df.head())
    
    # Select columns for visualization
    numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns.tolist()
    if not numeric_cols:
        st.warning("No numeric columns found for visualization.")
        return
    
    # Visualization type
    viz_type = st.selectbox("Select visualization type:", 
                           ["Bar Chart", "Line Chart", "Scatter Plot", "Histogram", "Box Plot", "Heatmap"])
    
    if viz_type in ["Bar Chart", "Line Chart", "Histogram", "Box Plot"]:
        col = st.selectbox("Select column:", numeric_cols)
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if viz_type == "Bar Chart":
            df[col].value_counts().plot(kind='bar', ax=ax)
        elif viz_type == "Line Chart":
            df[col].plot(kind='line', ax=ax)
        elif viz_type == "Histogram":
            df[col].plot(kind='hist', ax=ax)
        elif viz_type == "Box Plot":
            df[col].plot(kind='box', ax=ax)
            
        plt.title(f"{viz_type} of {col}")
        plt.tight_layout()
        st.pyplot(fig)
        
    elif viz_type == "Scatter Plot":
        col_x = st.selectbox("Select X-axis column:", numeric_cols)
        col_y = st.selectbox("Select Y-axis column:", [c for c in numeric_cols if c != col_x])
        
        fig, ax = plt.subplots(figsize=(10, 6))
        df.plot(kind='scatter', x=col_x, y=col_y, ax=ax)
        plt.title(f"Scatter Plot of {col_x} vs {col_y}")
        plt.tight_layout()
        st.pyplot(fig)
        
    elif viz_type == "Heatmap":
        if len(numeric_cols) < 2:
            st.warning("Need at least 2 numeric columns for a heatmap.")
            return
            
        corr = df[numeric_cols].corr()
        fig, ax = plt.subplots(figsize=(10, 8))
        sns.heatmap(corr, annot=True, cmap='coolwarm', ax=ax)
        plt.title("Correlation Heatmap")
        plt.tight_layout()
        st.pyplot(fig)

def handle_userinput(user_question):
    response = st.session_state.conversation({'question': user_question})
    st.session_state.chat_history = response['chat_history']

    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(user_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
        else:
            st.write(bot_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)

# Add these functions for model management

def get_local_ollama_models():
    """Get list of locally available Ollama models"""
    try:
        result = subprocess.run(['ollama', 'list'], 
                               stdout=subprocess.PIPE, 
                               stderr=subprocess.PIPE,
                               text=True,
                               check=True)
        
        # Parse the output to extract model names
        models = []
        for line in result.stdout.strip().split('\n')[1:]:  # Skip header line
            if line.strip():
                # Extract model name (first column)
                model_name = line.split()[0]
                models.append(model_name)
        
        return models
    except FileNotFoundError:
        st.error("Ollama not found. Please install Ollama to use local models.")
        return []
    except subprocess.CalledProcessError as e:
        st.error(f"Error getting Ollama models: {e}")
        return []

def pull_ollama_model(model_name):
    """Pull a new Ollama model"""
    try:
        with st.spinner(f"Pulling model {model_name}... This may take a while."):
            result = subprocess.run(['ollama', 'pull', model_name], 
                                   stdout=subprocess.PIPE, 
                                   stderr=subprocess.PIPE,
                                   text=True,
                                   check=True)
            st.success(f"Successfully pulled model: {model_name}")
            return True
    except subprocess.CalledProcessError as e:
        st.error(f"Error pulling model: {e}")
        return False

def get_popular_huggingface_models():
    """Get popular HuggingFace models for text generation"""
    # Default models in case API call fails
    default_models = [
        {"id": "google/flan-t5-xxl", "name": "Flan-T5-XXL (Large)"},
        {"id": "google/flan-t5-xl", "name": "Flan-T5-XL (Medium)"},
        {"id": "google/flan-t5-large", "name": "Flan-T5-Large (Small)"},
        {"id": "facebook/bart-large-cnn", "name": "BART Large CNN (Summarization)"}
    ]
    
    try:
        # Try to get popular models from HuggingFace API
        response = requests.get(
            "https://huggingface.co/api/models?sort=downloads&direction=-1&limit=100&filter=text-generation",
            timeout=5
        )
        
        if response.status_code == 200:
            all_models = response.json()
            # Filter for models that are likely to work well with LangChain
            compatible_models = []
            for model in all_models:
                # Look for popular, smaller models that are more likely to work
                if any(name in model['id'].lower() for name in ['flan', 't5', 'bart', 'gpt2', 'bloom']):
                    compatible_models.append({
                        "id": model['id'],
                        "name": model['id'].split('/')[-1]
                    })
                    if len(compatible_models) >= 10:
                        break
            
            return compatible_models if compatible_models else default_models
        else:
            return default_models
    except Exception as e:
        st.warning(f"Could not fetch HuggingFace models: {e}")
        return default_models

def get_openai_models():
    """Get available OpenAI models"""
    # Default list of models
    models = [
        {"id": "gpt-3.5-turbo", "name": "GPT-3.5 Turbo"},
        {"id": "gpt-4", "name": "GPT-4"},
        {"id": "gpt-4-turbo", "name": "GPT-4 Turbo"},
        {"id": "gpt-3.5-turbo-16k", "name": "GPT-3.5 Turbo (16K)"}
    ]
    
    # If we have an API key, try to get the actual list
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        try:
            headers = {
                "Authorization": f"Bearer {api_key}"
            }
            response = requests.get(
                "https://api.openai.com/v1/models",
                headers=headers,
                timeout=5
            )
            
            if response.status_code == 200:
                all_models = response.json()["data"]
                chat_models = []
                
                for model in all_models:
                    # Filter for chat models
                    if any(name in model['id'] for name in ['gpt-3.5', 'gpt-4']):
                        chat_models.append({
                            "id": model['id'],
                            "name": model['id']
                        })
                
                return chat_models if chat_models else models
        except Exception as e:
            st.warning(f"Could not fetch OpenAI models: {e}")
    
    return models

def update_env_file(key, value):
    """Update a key in the .env file"""
    env_path = Path('.env')
    
    # Create .env file if it doesn't exist
    if not env_path.exists():
        with open(env_path, 'w') as f:
            f.write(f"{key}={value}\n")
        return True
    
    # Read existing content
    with open(env_path, 'r') as f:
        lines = f.readlines()
    
    # Check if key exists and update it
    key_exists = False
    new_lines = []
    for line in lines:
        if line.strip() and line.split('=')[0].strip() == key:
            new_lines.append(f"{key}={value}\n")
            key_exists = True
        else:
            new_lines.append(line)
    
    # Add key if it doesn't exist
    if not key_exists:
        new_lines.append(f"{key}={value}\n")
    
    # Write back to file
    with open(env_path, 'w') as f:
        f.writelines(new_lines)
    
    # Also update the environment variable in the current session
    os.environ[key] = value
    
    return True

def show_embedding_model_info():
    """Display information about embedding models vs LLM models"""
    with st.sidebar.expander("ℹ️ About Embedding Models"):
        st.markdown("""
        **Note:** This application uses two types of AI models:
        
        1. **Embedding Model** - Always uses `all-MiniLM-L6-v2` to convert text to vector representations
        2. **Language Model** - Uses your selected provider (Ollama, HuggingFace, or OpenAI)
        
        The embedding model message you see is normal and doesn't affect your provider selection.
        """)

# Update the main function to include the new UI elements
def main():
    load_dotenv()
    st.set_page_config(page_title="Document Analysis Assistant",
                       page_icon="📚",
                       layout="wide")
    st.write(css, unsafe_allow_html=True)

    # Initialize session state variables
    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None
    if "text_chunks" not in st.session_state:
        st.session_state.text_chunks = None
    if "dataframes" not in st.session_state:
        st.session_state.dataframes = {}
    if "provider" not in st.session_state:
        st.session_state.provider = "HuggingFace"
    if "selected_model" not in st.session_state:
        st.session_state.selected_model = "google/flan-t5-xxl"
    if "ollama_models" not in st.session_state:
        st.session_state.ollama_models = []
    if "huggingface_models" not in st.session_state:
        st.session_state.huggingface_models = get_popular_huggingface_models()
    if "openai_models" not in st.session_state:
        st.session_state.openai_models = get_openai_models()
    if "show_new_model_input" not in st.session_state:
        st.session_state.show_new_model_input = False

    st.header("📚 Document Analysis Assistant")
    
    # Create tabs for different functionalities
    tabs = st.tabs(["Chat", "Summarize", "Visualize"])
    
    with tabs[0]:  # Chat tab
        st.subheader("Chat with your documents")
        user_question = st.text_input("Ask a question about your documents:")
        if user_question:
            if st.session_state.conversation is not None:
                handle_userinput(user_question)
            else:
                st.warning("Please upload and process documents first.")
    
    with tabs[1]:  # Summarize tab
        st.subheader("Generate document summary")
        if st.button("Generate Summary"):
            if st.session_state.text_chunks:
                with st.spinner("Generating summary..."):
                    use_ollama = st.session_state.provider == "Ollama"
                    use_openai = st.session_state.provider == "OpenAI"
                    summary = generate_summary(st.session_state.text_chunks, use_ollama, use_openai, model_name=st.session_state.selected_model)
                    st.write(summary)
            else:
                st.warning("Please upload and process documents first.")
    
    with tabs[2]:  # Visualize tab
        if st.session_state.dataframes:
            visualize_data(st.session_state.dataframes)
        else:
            st.info("Upload CSV or Excel files to enable visualization.")

    with st.sidebar:
        st.subheader("Document Settings")
        
        # Enhanced Model Selection
        st.subheader("Model Selection")
        
        # Add the embedding model info
        show_embedding_model_info()
        
        # Provider selection dropdown
        provider = st.selectbox(
            "Select AI Provider:",
            ["Ollama", "HuggingFace", "OpenAI"],
            index=["Ollama", "HuggingFace", "OpenAI"].index(st.session_state.provider)
        )
        
        # Update provider in session state
        if provider != st.session_state.provider:
            st.session_state.provider = provider
            # Reset selected model when changing provider
            st.session_state.selected_model = None
        
        # Show different model options based on provider
        if provider == "Ollama":
            # Refresh Ollama models
            if not st.session_state.ollama_models or st.button("Refresh Models"):
                with st.spinner("Getting local Ollama models..."):
                    st.session_state.ollama_models = get_local_ollama_models()
            
            # Show available models
            if st.session_state.ollama_models:
                model_name = st.selectbox(
                    "Select Ollama Model:",
                    st.session_state.ollama_models
                )
                st.session_state.selected_model = model_name
            else:
                st.warning("No Ollama models found. Please install Ollama and pull models.")
            
            # Option to pull new models
            col1, col2 = st.columns([3, 1])
            with col1:
                if st.button("+ Add New Model"):
                    st.session_state.show_new_model_input = True
            
            if st.session_state.show_new_model_input:
                with st.form("pull_model_form"):
                    new_model = st.text_input("Enter model name to pull (e.g., llama2, mistral):")
                    submitted = st.form_submit_button("Pull Model")
                    if submitted and new_model:
                        success = pull_ollama_model(new_model)
                        if success:
                            st.session_state.ollama_models = get_local_ollama_models()
                            st.session_state.show_new_model_input = False
                            st.experimental_rerun()
        
        elif provider == "HuggingFace":
            # API Key input
            hf_api_key = st.text_input(
                "HuggingFace API Key (optional):",
                type="password",
                value=os.getenv("HUGGINGFACEHUB_API_TOKEN", "")
            )
            
            if hf_api_key and hf_api_key != os.getenv("HUGGINGFACEHUB_API_TOKEN", ""):
                if st.button("Save HuggingFace API Key"):
                    update_env_file("HUGGINGFACEHUB_API_TOKEN", hf_api_key)
                    st.success("API key saved to .env file!")
            
            # Model selection
            model_options = [{"id": model["id"], "name": f"{model['id']} ({model['name']})" if 'name' in model else model['id']} 
                            for model in st.session_state.huggingface_models]
            
            selected_model_name = st.selectbox(
                "Select HuggingFace Model:",
                options=[model["name"] for model in model_options],
                index=0 if not st.session_state.selected_model else next(
                    (i for i, model in enumerate(model_options) if model["id"] == st.session_state.selected_model), 
                    0
                )
            )
            
            # Get the model ID from the selected name
            selected_model_id = next(
                (model["id"] for model in model_options if model["name"] == selected_model_name),
                model_options[0]["id"] if model_options else "google/flan-t5-xxl"
            )
            
            st.session_state.selected_model = selected_model_id
        
        elif provider == "OpenAI":
            # API Key input
            openai_api_key = st.text_input(
                "OpenAI API Key:",
                type="password",
                value=os.getenv("OPENAI_API_KEY", "")
            )
            
            if openai_api_key and openai_api_key != os.getenv("OPENAI_API_KEY", ""):
                if st.button("Save OpenAI API Key"):
                    update_env_file("OPENAI_API_KEY", openai_api_key)
                    st.success("API key saved to .env file!")
                    # Refresh OpenAI models with the new key
                    st.session_state.openai_models = get_openai_models()
            
            # Model selection
            if not openai_api_key:
                st.warning("Please enter your OpenAI API key to use OpenAI models.")
            
            model_options = [{"id": model["id"], "name": model["name"]} 
                            for model in st.session_state.openai_models]
            
            selected_model_name = st.selectbox(
                "Select OpenAI Model:",
                options=[model["name"] for model in model_options],
                index=0 if not st.session_state.selected_model else next(
                    (i for i, model in enumerate(model_options) if model["id"] == st.session_state.selected_model), 
                    0
                )
            )
            
            # Get the model ID from the selected name
            selected_model_id = next(
                (model["id"] for model in model_options if model["name"] == selected_model_name),
                model_options[0]["id"] if model_options else "gpt-3.5-turbo"
            )
            
            st.session_state.selected_model = selected_model_id
        
        # Document upload
        st.subheader("Upload Documents")
        
        # Add information about supported file types
        st.info("Supported file types: PDF, DOCX, CSV, Excel, EPUB, DJVU")
        
        docs = st.file_uploader(
            "Upload your documents here and click on 'Process'", 
            accept_multiple_files=True,
            type=["pdf", "docx", "csv", "xlsx", "xls", "epub", "djvu"]
        )
        
        if st.button("Process Documents"):
            if docs:
                with st.spinner("Processing documents..."):
                    # Get document text and dataframes
                    raw_text, dataframes = get_document_text(docs)
                    
                    if not raw_text:
                        st.error("No text could be extracted from the uploaded documents.")
                        st.stop()
                        
                    st.session_state.dataframes = dataframes
                    
                    # Get text chunks
                    text_chunks = get_text_chunks(raw_text)
                    st.session_state.text_chunks = text_chunks
                    
                    # Create vector store
                    st.info(f"Creating vector store using embedding model (this is separate from your selected {st.session_state.provider} model)")
                    vectorstore = get_vectorstore(text_chunks)
                    
                    # Create conversation chain based on selected provider
                    use_ollama = st.session_state.provider == "Ollama"
                    use_openai = st.session_state.provider == "OpenAI"
                    
                    st.info(f"Setting up conversation chain using your selected {st.session_state.provider} model: {st.session_state.selected_model}")
                    st.session_state.conversation = get_conversation_chain(
                        vectorstore, 
                        use_ollama=use_ollama,
                        use_openai=use_openai,
                        model_name=st.session_state.selected_model
                    )
                    
                    st.success(f"Processed {len(docs)} documents successfully!")
                    
                    # Show document stats
                    st.write(f"Total text length: {len(raw_text)} characters")
                    st.write(f"Number of chunks: {len(text_chunks)}")
                    if dataframes:
                        st.write(f"Data files available for visualization: {len(dataframes)}")
            else:
                st.warning("Please upload documents first.")

if __name__ == '__main__':
    main()

